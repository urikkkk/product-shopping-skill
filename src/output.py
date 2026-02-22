"""Output writers — XLSX, CSV, and Google Sheets."""

from __future__ import annotations

import csv
import logging
import os
from pathlib import Path
from typing import Any

from src.enrichment.reviews import ProReview
from src.schema import Product
from src.scoring import ScoreBreakdown, W_BUILD, W_ERGO, W_REVIEW, W_VALUE

logger = logging.getLogger(__name__)


def _dim_info(profile: Any | None) -> list[tuple[str, str]]:
    """Return list of (name, display_name) from profile or keyboard defaults."""
    if profile is not None:
        return [(d.name, d.display_name) for d in profile.dimensions]
    return [
        ("ergonomics", "Ergo"),
        ("reviews", "Review"),
        ("value", "Value"),
        ("build", "Build"),
    ]


def write_xlsx(
    products: list[Product],
    top10: list[tuple[Product, ScoreBreakdown]],
    reviews: dict[str, list[ProReview]],
    output_dir: str = "output",
    filename: str = "ergonomic_keyboards.xlsx",
    profile: Any | None = None,
) -> str:
    """Write styled XLSX with All Products + Top 10 + Pro Reviews tabs."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    os.makedirs(output_dir, exist_ok=True)
    path = os.path.join(output_dir, filename)
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    gold_fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
    highlight_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    dims = _dim_info(profile)

    # ── Tab 1: All Products ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "All Products"
    headers = [
        "Title", "Brand", "Price ($)", "Rating", "Reviews", "Availability",
        "Store", "Category", "Layout", "Switch", "Connectivity",
        "Hot-Swap", "Programmable", "Ergonomic Features", "URL",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for i, p in enumerate(products, 2):
        row = [
            p.product_title, p.brand, p.price_usd, p.rating_avg, p.rating_count,
            p.availability, p.source_site, p.category, p.layout_size,
            p.switch_type, p.connectivity,
            "Yes" if p.hot_swappable else "No", p.programmable,
            p.ergonomic_features, p.product_url,
        ]
        for col, val in enumerate(row, 1):
            ws.cell(row=i, column=col, value=val)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(products) + 1}"

    # ── Tab 2: Top 10 ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Top 10")
    top_headers = (
        ["Rank", "Title", "Brand", "Price ($)", "Rating", "Score"]
        + [dn for _, dn in dims]
        + ["Layout", "Connectivity", "Why It Made the List"]
    )
    for col, h in enumerate(top_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = gold_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for rank, (p, s) in enumerate(top10, 1):
        revs = reviews.get(p.product_title, [])
        reason = revs[0].verdict if revs else "Strong ergonomic features and value"
        row = (
            [rank, p.product_title, p.brand, p.price_usd, p.rating_avg, s.total]
            + [s.dimensions.get(name, 0.0) for name, _ in dims]
            + [p.layout_size, p.connectivity, reason]
        )
        for col, val in enumerate(row, 1):
            cell = ws2.cell(row=rank + 1, column=col, value=val)
            if rank <= 3:
                cell.fill = highlight_fill

    for col in range(1, len(top_headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 18
    # "Why It Made the List" is last column
    ws2.column_dimensions[get_column_letter(len(top_headers))].width = 60

    # ── Tab 3: Pro Reviews ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("Pro Reviews")
    rev_headers = ["Product", "Source", "Pros", "Cons", "Verdict", "Best For", "Ergo Notes"]
    green_fill = PatternFill(start_color="1B7340", end_color="1B7340", fill_type="solid")
    for col, h in enumerate(rev_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = green_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    row_num = 2
    for title, revs in reviews.items():
        for r in revs:
            for col, val in enumerate(
                [title, r.source, r.pros, r.cons, r.verdict, r.best_for, r.ergo_notes], 1
            ):
                ws3.cell(row=row_num, column=col, value=val)
            row_num += 1

    for col in range(1, len(rev_headers) + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 25
    ws3.column_dimensions["C"].width = 50
    ws3.column_dimensions["D"].width = 50

    wb.save(path)
    logger.info("XLSX saved: %s", path)
    return path


def write_csv(
    products: list[Product],
    output_dir: str = "output",
    filename: str = "keyboards.csv",
    profile: Any | None = None,
) -> str:
    """Write flat CSV for the web app."""
    os.makedirs(output_dir, exist_ok=True)
    path = os.path.join(output_dir, filename)

    from src.scoring import score_product

    dims = _dim_info(profile)
    dim_score_fields = [f"{name}_score" for name, _ in dims]
    fields = Product.field_names() + ["score"] + dim_score_fields

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for p in products:
            s = score_product(p, profile=profile)
            row = p.to_dict()
            row["score"] = s.total
            for name, _ in dims:
                row[f"{name}_score"] = s.dimensions.get(name, 0.0)
            writer.writerow(row)

    logger.info("CSV saved: %s", path)
    return path


def write_google_sheets(
    products: list[Product],
    top10: list[tuple[Product, ScoreBreakdown]],
    reviews: dict[str, list[ProReview]],
    sheet_id: str | None = None,
    profile: Any | None = None,
) -> str:
    """Write to Google Sheets with 3 tabs: All Products, Top 10, Pro Reviews.

    Stores all collected products (up to 1000+) with scores.
    Set up credentials following cookbook/02-google-sheets-credentials.md.
    Returns the spreadsheet URL.
    """
    try:
        from google.oauth2.credentials import Credentials
        from google_auth_oauthlib.flow import InstalledAppFlow
        from googleapiclient.discovery import build
    except ImportError:
        logger.error(
            "Google Sheets dependencies not installed. "
            "Run: pip install 'keyboard-shopping-agent[google]'"
        )
        raise

    from src.scoring import score_product

    dims = _dim_info(profile)

    scopes = ["https://www.googleapis.com/auth/spreadsheets"]
    creds = None
    token_path = Path("token.json")
    creds_path = Path("credentials.json")

    if token_path.exists():
        creds = Credentials.from_authorized_user_file(str(token_path), scopes)

    if not creds or not creds.valid:
        if not creds_path.exists():
            raise FileNotFoundError(
                "credentials.json not found. Follow cookbook/02-google-sheets-credentials.md"
            )
        flow = InstalledAppFlow.from_client_secrets_file(str(creds_path), scopes)
        creds = flow.run_local_server(port=0)
        token_path.write_text(creds.to_json())

    service = build("sheets", "v4", credentials=creds)

    # ── Create or reuse spreadsheet ───────────────────────────────────────
    if sheet_id:
        spreadsheet_id = sheet_id
    else:
        spreadsheet = (
            service.spreadsheets()
            .create(
                body={
                    "properties": {"title": "Ergonomic Keyboard Research"},
                    "sheets": [
                        {"properties": {"title": "All Products", "index": 0}},
                        {"properties": {"title": "Top 10", "index": 1}},
                        {"properties": {"title": "Pro Reviews", "index": 2}},
                    ],
                }
            )
            .execute()
        )
        spreadsheet_id = spreadsheet["spreadsheetId"]

    url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}"
    logger.info("Google Sheet: %s", url)

    # When reusing an existing sheet, ensure tabs exist
    if sheet_id:
        existing = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
        existing_titles = {s["properties"]["title"] for s in existing["sheets"]}
        requests = []
        for tab_title in ("All Products", "Top 10", "Pro Reviews"):
            if tab_title not in existing_titles:
                requests.append(
                    {"addSheet": {"properties": {"title": tab_title}}}
                )
        if requests:
            service.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body={"requests": requests},
            ).execute()

    # ── Tab 1: All Products (with scores) ─────────────────────────────────
    dim_score_headers = [f"{dn} Score" for _, dn in dims]
    all_headers = (
        ["Title", "Brand", "Price ($)", "Rating", "Reviews", "Score"]
        + dim_score_headers
        + [
            "Availability", "Store", "Category", "Layout", "Switch",
            "Connectivity", "Hot-Swap", "Programmable", "Ergonomic Features", "URL",
        ]
    )
    all_rows = [all_headers]
    for p in products:
        s = score_product(p, profile=profile)
        all_rows.append(
            [p.product_title, p.brand, p.price_usd, p.rating_avg, p.rating_count, s.total]
            + [s.dimensions.get(name, 0.0) for name, _ in dims]
            + [
                p.availability, p.source_site, p.category, p.layout_size,
                p.switch_type, p.connectivity,
                "Yes" if p.hot_swappable else "No", p.programmable,
                p.ergonomic_features, p.product_url,
            ]
        )

    # ── Tab 2: Top 10 ─────────────────────────────────────────────────────
    top_headers = (
        ["Rank", "Title", "Brand", "Price ($)", "Rating", "Score"]
        + [dn for _, dn in dims]
        + ["Layout", "Connectivity", "Why It Made the List"]
    )
    top_rows = [top_headers]
    for rank, (p, s) in enumerate(top10, 1):
        revs = reviews.get(p.product_title, [])
        reason = revs[0].verdict if revs else "Strong ergonomic features and value"
        top_rows.append(
            [rank, p.product_title, p.brand, p.price_usd, p.rating_avg, s.total]
            + [s.dimensions.get(name, 0.0) for name, _ in dims]
            + [p.layout_size, p.connectivity, reason]
        )

    # ── Tab 3: Pro Reviews ────────────────────────────────────────────────
    rev_headers = ["Product", "Source", "Pros", "Cons", "Verdict", "Best For", "Ergo Notes"]
    rev_rows = [rev_headers]
    for title, revs in reviews.items():
        for r in revs:
            rev_rows.append([
                title, r.source, r.pros, r.cons, r.verdict, r.best_for, r.ergo_notes,
            ])

    # ── Batch write all 3 tabs ────────────────────────────────────────────
    service.spreadsheets().values().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={
            "valueInputOption": "RAW",
            "data": [
                {"range": "All Products!A1", "values": all_rows},
                {"range": "Top 10!A1", "values": top_rows},
                {"range": "Pro Reviews!A1", "values": rev_rows},
            ],
        },
    ).execute()

    # ── Format header rows (bold + freeze) ────────────────────────────────
    sheet_meta = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
    sheet_ids = {
        s["properties"]["title"]: s["properties"]["sheetId"]
        for s in sheet_meta["sheets"]
    }

    format_requests = []
    for tab_title in ("All Products", "Top 10", "Pro Reviews"):
        sid = sheet_ids.get(tab_title)
        if sid is None:
            continue
        # Bold header row
        format_requests.append({
            "repeatCell": {
                "range": {"sheetId": sid, "startRowIndex": 0, "endRowIndex": 1},
                "cell": {
                    "userEnteredFormat": {
                        "textFormat": {"bold": True},
                        "backgroundColor": {"red": 0.18, "green": 0.33, "blue": 0.59},
                        "textFormat": {"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
                    }
                },
                "fields": "userEnteredFormat(textFormat,backgroundColor)",
            }
        })
        # Freeze header row
        format_requests.append({
            "updateSheetProperties": {
                "properties": {"sheetId": sid, "gridProperties": {"frozenRowCount": 1}},
                "fields": "gridProperties.frozenRowCount",
            }
        })
        # Auto-resize columns
        format_requests.append({
            "autoResizeDimensions": {
                "dimensions": {"sheetId": sid, "dimension": "COLUMNS", "startIndex": 0, "endIndex": 20},
            }
        })

    if format_requests:
        service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": format_requests},
        ).execute()

    logger.info(
        "Google Sheet updated — %d products, %d top picks, %d reviews: %s",
        len(products), len(top10), len(rev_rows) - 1, url,
    )
    return url
